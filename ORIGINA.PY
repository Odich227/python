import sys
import openpyxl
from openpyxl import Workbook
import os
import re
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QLabel, 
    QLineEdit, QPushButton, QMessageBox, QFormLayout,
    QTableWidget, QTableWidgetItem, QTabWidget, QHeaderView,
    QComboBox, QDateEdit
)
from PyQt6.QtCore import Qt, QDate

# Константы
EXCEL_FILE = "users_registration.xlsx"
SHEET_NAME = "Users"

class RegistrationApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Система регистрации пользователей")
        self.setFixedSize(1000, 700)
        
        # Создаем вкладки
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        
        # Вкладка регистрации
        self.tab_register = QWidget()
        self.setup_register_tab()
        
        # Вкладка просмотра пользователей
        self.tab_view = QWidget()
        self.setup_view_tab()
        
        self.tabs.addTab(self.tab_register, "Регистрация")
        self.tabs.addTab(self.tab_view, "Просмотр пользователей")
        
        # Инициализация Excel файла
        self.init_excel_file()
    
    def setup_register_tab(self):
        """Настройка вкладки регистрации"""
        layout = QVBoxLayout()
        self.tab_register.setLayout(layout)
        
        # Форма для ввода данных
        form_layout = QFormLayout()
        
        # Основные поля
        self.label_username = QLabel("Логин*:")
        self.edit_username = QLineEdit()
        form_layout.addRow(self.label_username, self.edit_username)
        
        self.label_password = QLabel("Пароль*:")
        self.edit_password = QLineEdit()
        form_layout.addRow(self.label_password, self.edit_password)
        
        self.label_email = QLabel("Email*:")
        self.edit_email = QLineEdit()
        form_layout.addRow(self.label_email, self.edit_email)
        
        # Персональные данные
        self.label_lastname = QLabel("Фамилия:")
        self.edit_lastname = QLineEdit()
        form_layout.addRow(self.label_lastname, self.edit_lastname)
        
        self.label_firstname = QLabel("Имя*:")
        self.edit_firstname = QLineEdit()
        form_layout.addRow(self.label_firstname, self.edit_firstname)
        
        self.label_middlename = QLabel("Отчество:")
        self.edit_middlename = QLineEdit()
        form_layout.addRow(self.label_middlename, self.edit_middlename)
        
        self.label_birthdate = QLabel("Дата рождения:")
        self.edit_birthdate = QDateEdit()
        self.edit_birthdate.setDisplayFormat("dd.MM.yyyy")
        self.edit_birthdate.setMaximumDate(QDate.currentDate())
        form_layout.addRow(self.label_birthdate, self.edit_birthdate)
        
        self.label_phone = QLabel("Телефон:")
        self.edit_phone = QLineEdit()
        self.edit_phone.setPlaceholderText("+71234567890")
        form_layout.addRow(self.label_phone, self.edit_phone)
        
        self.label_gender = QLabel("Пол:")
        self.combo_gender = QComboBox()
        self.combo_gender.addItems(["", "Мужской", "Женский"])
        form_layout.addRow(self.label_gender, self.combo_gender)
        
        layout.addLayout(form_layout)
        
        # Кнопка регистрации
        self.button_register = QPushButton("Зарегистрироваться")
        self.button_register.clicked.connect(self.register_user)
        layout.addWidget(self.button_register)
    
    def setup_view_tab(self):
        """Настройка вкладки просмотра пользователей"""
        layout = QVBoxLayout()
        self.tab_view.setLayout(layout)
        
        # Таблица для отображения пользователей
        self.table_users = QTableWidget()
        self.table_users.setColumnCount(10)
        self.table_users.setHorizontalHeaderLabels([
            "ID", "Логин", "Email", "Фамилия", "Имя", "Отчество", 
            "Дата рождения", "Телефон", "Пол", "Дата регистрации"
        ])
        self.table_users.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        
        layout.addWidget(self.table_users)
        
        # Кнопка обновления таблицы
        self.button_refresh = QPushButton("Обновить список")
        self.button_refresh.clicked.connect(self.load_users_to_table)
        layout.addWidget(self.button_refresh)
    
    def init_excel_file(self):
        """Инициализация Excel файла"""
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append([
                "ID", "Username", "Password", "Email", 
                "Lastname", "Firstname", "Middlename",
                "Birthdate", "Phone", "Gender", "Registration Date"
            ])
            wb.save(EXCEL_FILE)
    
    def is_valid_email(self, email):
        """Проверка валидности email"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    def is_valid_phone(self, phone):
        """Проверка валидности телефона"""
        if not phone:
            return True
        pattern = r'^\+?[0-9\s\-\(\)]{7,20}$'
        return re.match(pattern, phone) is not None
    
    def get_next_id(self):
        """Получение следующего ID (максимальный + 1)"""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            max_id = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], int) and row[0] > max_id:
                    max_id = row[0]
            return max_id + 1
        except FileNotFoundError:
            return 1
    
    def check_existing_user(self, username, email):
        """Проверка существующего пользователя"""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == username:
                    return "Логин уже занят"
                if row[3] == email:
                    return "Email уже зарегистрирован"
                    
        except FileNotFoundError:
            pass
            
        return None
    
    def register_user(self):
        """Регистрация пользователя"""
        # Получаем данные из полей
        username = self.edit_username.text().strip()
        password = self.edit_password.text()
        email = self.edit_email.text().strip()
        lastname = self.edit_lastname.text().strip()
        firstname = self.edit_firstname.text().strip()
        middlename = self.edit_middlename.text().strip()
        birthdate = self.edit_birthdate.date().toString("dd.MM.yyyy")
        phone = self.edit_phone.text().strip()
        gender = self.combo_gender.currentText()
        
        # Валидация обязательных полей
        if not username:
            QMessageBox.warning(self, "Ошибка", "Поле 'Логин' обязательно для заполнения!")
            return
        
        if not password:
            QMessageBox.warning(self, "Ошибка", "Поле 'Пароль' обязательно для заполнения!")
            return
            
        if not email:
            QMessageBox.warning(self, "Ошибка", "Поле 'Email' обязательно для заполнения!")
            return
            
        if not firstname:
            QMessageBox.warning(self, "Ошибка", "Поле 'Имя' обязательно для заполнения!")
            return
            
        # Проверка email
        if not self.is_valid_email(email):
            QMessageBox.warning(self, "Ошибка", "Введите корректный email адрес!")
            return
            
        # Проверка телефона
        if not self.is_valid_phone(phone):
            QMessageBox.warning(self, "Ошибка", "Введите корректный номер телефона!")
            return
            
        # Проверка существующего пользователя
        existing_error = self.check_existing_user(username, email)
        if existing_error:
            QMessageBox.warning(self, "Ошибка", existing_error)
            return
        
        # Получаем следующий ID
        user_id = self.get_next_id()
        
        # Получаем текущую дату
        reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Записываем данные в Excel
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            ws.append([
                user_id, username, password, email, 
                lastname, firstname, middlename,
                birthdate, phone, gender, reg_date
            ])
            wb.save(EXCEL_FILE)
            
            QMessageBox.information(
                self, 
                "Успех", 
                f"Пользователь {username} зарегистрирован!\nID: {user_id}"
            )
            
            # Очищаем поля после регистрации
            self.edit_username.clear()
            self.edit_password.clear()
            self.edit_email.clear()
            self.edit_lastname.clear()
            self.edit_firstname.clear()
            self.edit_middlename.clear()
            self.edit_birthdate.setDate(QDate.currentDate())
            self.edit_phone.clear()
            self.combo_gender.setCurrentIndex(0)
            
            # Обновляем таблицу
            self.load_users_to_table()
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Ошибка", 
                f"Не удалось сохранить данные:\n{str(e)}"
            )
    
    def load_users_to_table(self):
        """Загрузка пользователей в таблицу"""
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb[SHEET_NAME]
            
            # Получаем все строки (начиная со 2-й)
            users_data = list(ws.iter_rows(min_row=2, values_only=True))
            
            # Настраиваем таблицу
            self.table_users.setRowCount(len(users_data))
            
            # Заполняем таблицу данными
            for row_idx, row_data in enumerate(users_data):
                for col_idx, cell_data in enumerate(row_data):
                    if col_idx == 2:  # Пароль - скрываем
                        item = QTableWidgetItem("******")
                    else:
                        item = QTableWidgetItem(str(cell_data if cell_data is not None else ""))
                    
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.table_users.setItem(row_idx, col_idx, item)
                    
            # Автоматическое изменение размера столбцов
            self.table_users.resizeColumnsToContents() 
        except Exception as e:
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось загрузить данные:\n{str(e)}"
            )

if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = RegistrationApp()
    window.show()
    sys.exit(app.exec())

    